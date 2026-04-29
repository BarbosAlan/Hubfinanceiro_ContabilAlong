from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Iterable, Optional
import re
import unicodedata

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class GenialTargets:
    payout: str = "PAY OUT"
    payin: str = "PAY IN"
    bloqueio: str = "BLOQUEIO"
    desbloqueio: str = "DESBLOQUEIO"

    @property
    def all(self) -> list[str]:
        return [self.payout, self.payin, self.bloqueio, self.desbloqueio]


REQUIRED_COLUMNS = ["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"]


def _norm_text(s: object) -> str:
    """
    Normaliza texto para matching:
    - upper
    - remove acentos
    - remove pontuação simples
    - colapsa espaços
    """
    txt = "" if s is None else str(s)
    txt = unicodedata.normalize("NFKD", txt)
    txt = "".join(ch for ch in txt if not unicodedata.combining(ch))
    txt = txt.upper()
    txt = re.sub(r"[^A-Z0-9 ]+", " ", txt)
    txt = re.sub(r"\s+", " ", txt).strip()
    return txt


def classify_histórico(value: object, *, targets: GenialTargets | None = None) -> Optional[str]:
    """
    Classifica o histórico por palavras-chave (contém), tolerante a variações:
    - "PAGAMENTO DE PAY OUT - BILL LINE" -> "PAY OUT"
    - "RECEBIMENTO DE PAY IN ..." -> "PAY IN"
    - variações com/sem espaço: "PAYOUT", "PAYIN"
    - bloqueio/desbloqueio (prioriza DESBLOQUEIO quando ambos aparecem)
    """
    t = targets or GenialTargets()
    s = _norm_text(value)
    if not s:
        return None

    # desbloqueio primeiro (para não cair em bloqueio)
    if "DESBLOQ" in s or ("DES" in s and "BLOQ" in s):
        return t.desbloqueio

    if "BLOQ" in s:
        return t.bloqueio

    # PAY IN/OUT (aceita "PAYIN"/"PAYOUT")
    if "PAY" in s:
        s2 = s.replace(" ", "")
        if "PAYOUT" in s2 or ("PAY" in s and "OUT" in s):
            return t.payout
        if "PAYIN" in s2 or ("PAY" in s and re.search(r"\bIN\b", s)):
            return t.payin

    return None


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    renames: dict[str, str] = {}

    # Fuzzy detection: order/case/acentos não importam.
    # Preferência: primeiro detecta "histórico de lançamento", depois histórico simples.
    found = {
        "Data": None,
        "Valor": None,
        "HISTORICO": None,
        "HISTORICO DE LANÇAMENTO": None,
    }

    for orig in df.columns:
        n = _norm_text(orig)
        if not n:
            continue

        if found["Data"] is None and ("DATA" in n or "DATE" in n):
            found["Data"] = orig
            continue

        if found["Valor"] is None and ("VALOR" in n or "VALUE" in n or "AMOUNT" in n):
            found["Valor"] = orig
            continue

        if "HIST" in n:
            if found["HISTORICO DE LANÇAMENTO"] is None and ("LANC" in n or "LAN" in n):
                found["HISTORICO DE LANÇAMENTO"] = orig
                continue
            if found["HISTORICO"] is None:
                found["HISTORICO"] = orig
                continue

    for target, orig in found.items():
        if orig is not None:
            renames[orig] = target

    return df.rename(columns=renames)


def process_genial_excel(
    excel_bytes: bytes,
    *,
    targets: Iterable[str] | None = None,
) -> tuple[pd.DataFrame, dict]:
    """
    Processa o Excel do Extrato Genial de forma dinâmica.
    - Identifica por palavras-chave (PAY IN, PAY OUT, BLOQUEIO, etc.)
    - Agrupa por Data e Tipo.
    - Preserva as frases originais do arquivo.
    """
    df = pd.read_excel(BytesIO(excel_bytes))
    df = _normalize_columns(df)

    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Colunas faltando no arquivo: {', '.join(missing)}")

    df["HISTORICO"] = df["HISTORICO"].astype(str).str.strip()
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce").fillna(0.0)
    df["Data"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
    linhas_data_invalida = int(df["Data"].isna().sum())
    df = df.dropna(subset=["Data"])
    df["Data"] = df["Data"].dt.date

    klass_targets = GenialTargets()
    allowed = set(list(targets)) if targets is not None else set(klass_targets.all)

    # Classifica as linhas
    df["_CLASS"] = df["HISTORICO"].apply(lambda v: classify_histórico(v, targets=klass_targets))

    df_targets = df[df["_CLASS"].isin(allowed)].copy()
    df_others = df[~df["_CLASS"].isin(allowed)].copy()

    resultados = []
    
    # Agrupamento Dinâmico para linhas com palavras-chave
    if not df_targets.empty:
        # Agrupamos por Data e pela Classe encontrada (PAY IN, PAY OUT, etc.)
        grouped = df_targets.groupby(["Data", "_CLASS"], sort=False)
        for (data, klass), grupo in grouped:
            # 1. Recupera as frases originais únicas do Histórico
            h_unicos = list(dict.fromkeys(grupo['HISTORICO'].astype(str).tolist()))
            txt_h = " | ".join(h_unicos) if len(h_unicos) > 1 else h_unicos[0]
            
            # 2. Agora, conforme solicitado, apenas copia o nome que ficou no histórico
            # (Sem juntar detalhes ou criar listas numeradas)
            txt_l = txt_h

            resultados.append({
                "Data": data,
                "HISTORICO": txt_h,
                "Valor": grupo['Valor'].sum(),
                "HISTORICO DE LANÇAMENTO": txt_l
            })

    # Linhas sem palavras-chave: mantém ambas as colunas com valor original do arquivo
    for _, row in df_others.iterrows():
        resultados.append({
            "Data": row["Data"],
            "HISTORICO": row["HISTORICO"],
            "Valor": row["Valor"],
            "HISTORICO DE LANÇAMENTO": row["HISTORICO DE LANÇAMENTO"]
        })

    if not resultados:
        df_final = pd.DataFrame(columns=["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"])
    else:
        df_final = pd.DataFrame(resultados)
        df_final = df_final.sort_values(["Data", "HISTORICO"]).reset_index(drop=True)

    stats = {
        "linhas_originais": int(len(df)),
        "outros_mantidos": int(len(df_others)),
        "agrupados": int(len(df_targets)),
        "total_saida": int(len(df_final)),
        "saldo_final": round(float(df_final["Valor"].sum()), 2) if not df_final.empty else 0.0,
        "datas_invalidas": linhas_data_invalida,
    }

    return df_final, stats


def format_genial_excel(df_final: pd.DataFrame, *, targets: Iterable[str] | None = None) -> BytesIO:
    """
    Gera XLSX formatado (equivalente ao `formatar_excel` do `app.py`).
    Retorna um BytesIO pronto para download.
    """
    targets_set = set(list(targets) if targets is not None else GenialTargets().all)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato Tratado"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    headers = ["Data", "HISTORICO", "Valor", "HISTORICO DE LANÇAMENTO"]
    col_widths = [15, 45, 18, 45]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22

    for idx, row in df_final.iterrows():
        r = idx + 2
        hist = str(row["HISTORICO"]).strip()

        dc = ws.cell(row=r, column=1, value=row["Data"])
        dc.number_format = "DD/MM/YYYY"
        dc.alignment = center
        dc.border = border

        hc = ws.cell(row=r, column=2, value=hist)
        hc.alignment = left_align
        hc.border = border
        hc.font = Font(name="Arial", size=10)

        vc = ws.cell(row=r, column=3, value=row["Valor"])
        vc.number_format = "#,##0.00"
        vc.alignment = center
        vc.border = border

        if classify_histórico(hist) is not None:
            cor = "C00000" if float(row["Valor"]) < 0 else "375623"
            vc.font = Font(name="Arial", color=cor, size=10, bold=True)
        else:
            vc.font = Font(name="Arial", size=10)

        hlc = ws.cell(row=r, column=4, value=row["HISTORICO DE LANÇAMENTO"])
        hlc.alignment = left_align
        hlc.border = border
        hlc.font = Font(name="Arial", size=10)

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

